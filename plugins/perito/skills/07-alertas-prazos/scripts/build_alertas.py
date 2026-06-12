#!/usr/bin/env python3
"""
build_alertas.py — LEITOR (read-only) da planilha de planejamento do perito Irineu.
NÃO escreve, NÃO altera a planilha. Só lê e monta o texto do planejamento da semana.

Uso:
    python3 build_alertas.py <planilha.xlsx> [--hoje AAAA-MM-DD] [--dias 15]

Saída: imprime o corpo do e-mail/planejamento em texto (stdout).
Foco: diligências a fazer · laudos a entregar · impugnações a responder (próx. N dias),
      com ATRASADOS em destaque no topo, prazos de AGENDAMENTO e diligências agrupadas por dia+vara.
Coluna que marca "laudo já entregue" = ENTREGA LAUDO (data preenchida) → sai do radar.
"""
import sys, re, datetime as dt

# --- auto-provisionamento de dependências (sandbox efêmero do Cowork) ---
def _ensure(pkgs):
    import importlib.util, subprocess, sys as _sys
    falta = [pip for mod, pip in pkgs if importlib.util.find_spec(mod) is None]
    if falta:
        cmd = [_sys.executable, '-m', 'pip', 'install', *falta]
        if _sys.platform.startswith('linux'):
            cmd.append('--break-system-packages')
        subprocess.run(cmd, check=False)
_ensure([('openpyxl', 'openpyxl')])

import openpyxl

CNJ = re.compile(r'\d{7}-\d{2}\.\d{4}\.\d')

def empty(v):
    if v is None: return True
    if isinstance(v, str) and v.strip().lower() in ('', 'o', '-', '#value!'): return True
    return False

def asdate(v):
    if isinstance(v, dt.datetime): return v.date()
    if isinstance(v, dt.date): return v
    return None

def hhmm(v):
    if isinstance(v, dt.datetime): return v.strftime('%H:%M')
    if isinstance(v, dt.time): return v.strftime('%H:%M')
    return (str(v)[:5] if not empty(v) else '')

def colmap(ws, header_row=1, ncols=30):
    m = {}
    for c in range(1, ncols+1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            m[v.strip().upper()] = c
    return m

def find(m, *names):
    for n in names:
        if n.upper() in m: return m[n.upper()]
    return None

def read_trabalhista(wb):
    ws = wb['Trabalhista']
    m = colmap(ws)
    C = dict(
        proc=find(m,'PROCESSO'), vara=find(m,'VARA'), rte=find(m,'RECLAMANTE'), rda=find(m,'RECLAMADA'),
        agdia=find(m,'AGENDA (DIA)'), aghora=find(m,'AGENDA (HORA)'), agate=find(m,'AGENDA (ATÉ)'),
        dfinal=find(m,'DATA FINAL'), entrega=find(m,'ENTREGA LAUDO'),
        impug=find(m,'IMPUGNAÇÕES'),  # col "IMPUGNAÇÕES" = prazo do PERITO responder (≈5d após o prazo das partes)
        conclus=find(m,'CONCLUS.'),
    )
    rows = []
    for r in range(2, ws.max_row+1):
        p = ws.cell(r, C['proc']).value if C['proc'] else None
        if not (isinstance(p, str) and CNJ.search(p)): continue
        g = lambda k: ws.cell(r, C[k]).value if C[k] else None
        rows.append(dict(
            proc=p.strip(), vara=g('vara'), rte=g('rte'), rda=g('rda'),
            agdia=asdate(g('agdia')), aghora=g('aghora'), agate=asdate(g('agate')),
            dfinal=asdate(g('dfinal')), entrega=asdate(g('entrega')), impug=asdate(g('impug')),
            entrega_raw=g('entrega'),
        ))
    return rows

def lbl(x, n):
    quando = f"ATRASADO há {abs(n)}d" if n < 0 else (f"vence em {n}d" if n > 0 else "vence HOJE")
    vara = str(x['vara']).strip() if not empty(x['vara']) else '—'
    rte = str(x['rte']).strip() if not empty(x['rte']) else '—'
    return f"   • {x['proc']} · {vara} · {rte} · {quando}"

def build(planilha, hoje, dias, atraso_max=10):
    wb = openpyxl.load_workbook(planilha, data_only=True)
    rows = read_trabalhista(wb)
    horizon = hoje + dt.timedelta(days=dias)
    entregue = lambda x: not empty(x['entrega_raw'])          # ENTREGA LAUDO preenchida = já entregue
    agendado = lambda x: x['agdia'] is not None               # já tem diligência marcada
    # atrasado entra no topo só se venceu nos últimos `atraso_max` dias (evita prazos antigos virarem ruído)
    recente = lambda n: -atraso_max <= n < 0

    # coletas por categoria (dentro da janela, ainda pendentes)
    laudos, impugs, agend = [], [], []
    atrasados = []
    for x in rows:
        # LAUDOS A ENTREGAR (prazo do juiz = DATA FINAL), ainda não entregues
        if x['dfinal'] and not entregue(x):
            n = (x['dfinal'] - hoje).days
            if recente(n): atrasados.append(('Laudo', n, x, x['dfinal']))
            elif 0 <= n <= dias: laudos.append((n, x))
        # IMPUGNAÇÕES — prazo do perito responder (col IMPUGNAÇÕES)
        if x['impug']:
            n = (x['impug'] - hoje).days
            if recente(n): atrasados.append(('Impugnação', n, x, x['impug']))
            elif 0 <= n <= dias: impugs.append((n, x))
        # PRAZOS DE AGENDAMENTO (ainda não marcado)
        if x['agate'] and not agendado(x):
            n = (x['agate'] - hoje).days
            if recente(n): atrasados.append(('Agendar', n, x, x['agate']))
            elif 0 <= n <= dias: agend.append((n, x))

    # DILIGÊNCIAS MARCADAS na janela, agrupadas por dia + vara
    dilig = [x for x in rows if x['agdia'] and 0 <= (x['agdia']-hoje).days <= dias]
    dilig.sort(key=lambda x: (x['agdia'], hhmm(x['aghora'])))

    out = []
    out.append(f"PLANEJAMENTO DA SEMANA — Eng. Irineu de Freitas Branco Junior")
    out.append(f"Gerado em {hoje.strftime('%d/%m/%Y')} · horizonte: próximos {dias} dias (até {horizon.strftime('%d/%m/%Y')})")
    out.append("")

    # ATRASADOS NO TOPO
    atrasados.sort(key=lambda t: t[1])
    out.append(f"🔴 ATRASADOS ({len(atrasados)}) — venceram nos últimos {atraso_max} dias · resolver com prioridade")
    if not atrasados: out.append("   (nada atrasado ✅)")
    for tipo, n, x, d in atrasados:
        vara = str(x['vara']).strip() if not empty(x['vara']) else '—'
        rte = str(x['rte']).strip() if not empty(x['rte']) else '—'
        out.append(f"   • [{tipo}] {x['proc']} · {vara} · {rte} · há {abs(n)}d ({d.strftime('%d/%m')})")
    out.append("")

    # DILIGÊNCIAS POR DIA + VARA
    out.append(f"🔍 DILIGÊNCIAS A FAZER ({len(dilig)}) — agrupadas por dia e vara")
    if not dilig: out.append("   (nenhuma marcada na janela)")
    cur = None
    for x in dilig:
        if x['agdia'] != cur:
            cur = x['agdia']
            out.append(f"   {cur.strftime('%a %d/%m')}:")
        vara = str(x['vara']).strip() if not empty(x['vara']) else '—'
        rte = str(x['rte']).strip() if not empty(x['rte']) else '—'
        out.append(f"      - {hhmm(x['aghora']) or '--:--'} · {vara} · {x['proc']} · {rte}")
    out.append("")

    # LAUDOS A ENTREGAR
    laudos.sort(key=lambda t: t[0])
    out.append(f"📄 LAUDOS A ENTREGAR ({len(laudos)}) — por prazo do juiz (Data Final)")
    if not laudos: out.append("   (nenhum na janela)")
    for n, x in laudos: out.append(lbl(x, n) + f" ({x['dfinal'].strftime('%d/%m')})")
    out.append("")

    # IMPUGNAÇÕES
    impugs.sort(key=lambda t: t[0])
    out.append(f"✋ IMPUGNAÇÕES A RESPONDER ({len(impugs)}) — prazo de resposta do perito · conferir no PJE se houve impugnação")
    if not impugs: out.append("   (nenhuma na janela)")
    for n, x in impugs: out.append(lbl(x, n) + f" ({x['impug'].strftime('%d/%m')})")
    out.append("")

    # PRAZOS DE AGENDAMENTO
    agend.sort(key=lambda t: t[0])
    out.append(f"📌 PRAZOS DE AGENDAMENTO ({len(agend)}) — perícias a MARCAR (Agenda até)")
    if not agend: out.append("   (nenhum na janela)")
    for n, x in agend: out.append(lbl(x, n) + f" ({x['agate'].strftime('%d/%m')})")
    out.append("")
    out.append("— Fim. Planilha lida em modo somente-leitura; nada foi alterado. —")
    return "\n".join(out)

def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    planilha = sys.argv[1]
    hoje = dt.date.today(); dias = 15; atraso = 10
    if '--hoje' in sys.argv:
        hoje = dt.date.fromisoformat(sys.argv[sys.argv.index('--hoje')+1])
    if '--dias' in sys.argv:
        dias = int(sys.argv[sys.argv.index('--dias')+1])
    if '--atraso' in sys.argv:
        atraso = int(sys.argv[sys.argv.index('--atraso')+1])
    print(build(planilha, hoje, dias, atraso))

if __name__ == '__main__':
    main()
